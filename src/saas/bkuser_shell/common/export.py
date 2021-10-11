# -*- coding: utf-8 -*-
"""
Tencent is pleased to support the open source community by making 蓝鲸智云PaaS平台社区版 (BlueKing PaaS Community
Edition) available.
Copyright (C) 2017-2020 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the MIT License (the "License"); you may not use this file except in compliance with the License.
You may obtain a copy of the License at
http://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
specific language governing permissions and limitations under the License.
"""
import logging
from dataclasses import dataclass
from typing import TYPE_CHECKING, List

from bkuser_shell.config_center.constants import DynamicFieldTypeEnum
from bkuser_shell.organization.serializers.profiles import ProfileExportSerializer
from bkuser_shell.organization.utils import get_options_values_by_key
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, colors

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook


logger = logging.getLogger(__name__)


@dataclass
class ProfileExcelExporter:
    """导出"""

    workbook: "Workbook"
    exported_file_name: str
    fields: list
    title_row_index: int = 2

    def __post_init__(self):
        self.first_sheet = self.workbook.worksheets[0]
        self.first_sheet.alignment = Alignment(wrapText=True)

    def update_profiles(self, profiles: List[dict], extra_infos: dict = None):
        self._update_sheet_titles()

        for p_index, p in enumerate(profiles):

            exported_profile = ProfileExportSerializer(p).data
            for f_index, f in enumerate(self.fields):
                field_name = f["name"]

                try:
                    if f["builtin"]:
                        raw_value = exported_profile[field_name]
                    else:
                        raw_value = exported_profile["extras"][field_name]
                except KeyError:
                    # 当无法从当前用户属性中找到对应字段时，尝试从 extra_infos 中获取
                    if extra_infos is None:
                        logger.exception("failed to get value from field<%s>", f)
                        continue

                    try:
                        raw_value = extra_infos[str(p["id"])][field_name]
                    except KeyError:
                        logger.exception("failed to get value from field<%s>", f)
                        continue

                value = raw_value
                # options 存储值为 key， 但是 Excel 交互值为 value
                if f["type"] == DynamicFieldTypeEnum.ONE_ENUM.value:
                    value = ",".join(get_options_values_by_key(f["options"], [raw_value]))
                elif f["type"] == DynamicFieldTypeEnum.MULTI_ENUM.value:
                    value = ",".join(get_options_values_by_key(f["options"], raw_value))

                # 为电话添加国际号码段
                if f["name"] == "telephone":
                    value = f'+{exported_profile["country_code"]}{exported_profile[field_name]}'

                if raw_value is None:
                    continue

                self.first_sheet.cell(row=p_index + self.title_row_index + 1, column=f_index + 1, value=value)

    def _update_sheet_titles(self):
        """更新表格标题"""
        required_field_names = [x["display_name"] for x in self.fields if x["builtin"]]
        not_required_field_names = [x["display_name"] for x in self.fields if not x["builtin"]]

        red_ft = Font(color=colors.COLOR_INDEX[2])
        black_ft = Font(color=colors.BLACK)
        for index, field_name in enumerate(required_field_names):
            _cell = self.first_sheet.cell(
                row=self.title_row_index,
                column=index + 1,
                value=field_name,
            )
            _cell.font = red_ft

        for index, field_name in enumerate(not_required_field_names):
            _cell = self.first_sheet.cell(
                row=self.title_row_index,
                column=index + 1 + len(required_field_names),
                value=field_name,
            )
            _cell.font = black_ft

    def to_response(self) -> HttpResponse:
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f"attachment;filename={self.exported_file_name}.xlsx"
        self.workbook.save(response)
        return response

# -*- coding: utf-8 -*-
"""
TencentBlueKing is pleased to support the open source community by making 蓝鲸智云-用户管理(Bk-User) available.
Copyright (C) 2017-2021 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the MIT License (the "License"); you may not use this file except in compliance with the License.
You may obtain a copy of the License at http://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
specific language governing permissions and limitations under the License.
"""
import logging
from typing import List

from django.conf import settings
from django.db.models import Q
from django.utils import translation
from openpyxl import load_workbook
from rest_framework import generics, status
from rest_framework.parsers import FileUploadParser
from rest_framework.response import Response

from .serializers import (
    CategoryCreateInputSLZ,
    CategoryDepartmentListInputSLZ,
    CategoryDetailOutputSLZ,
    CategoryExportInputSLZ,
    CategoryExportProfileOutputSLZ,
    CategoryFileImportInputSLZ,
    CategoryFileImportQuerySLZ,
    CategoryMetaOutputSLZ,
    CategoryNamespaceSettingUpdateInputSLZ,
    CategoryProfileListInputSLZ,
    CategorySettingCreateInputSLZ,
    CategorySettingOutputSLZ,
    CategorySyncResponseOutputSLZ,
    CategoryTestConnectionInputSLZ,
    CategoryTestFetchDataInputSLZ,
    CategoryUpdateInputSLZ,
)
from bkuser_core.api.web.department.serializers import DepartmentsWithChildrenAndAncestorsOutputSLZ
from bkuser_core.api.web.export import ProfileExcelExporter
from bkuser_core.api.web.field.serializers import FieldOutputSLZ
from bkuser_core.api.web.serializers import ProfileDetailOutputSLZ
from bkuser_core.api.web.utils import get_category, get_operator, list_setting_metas
from bkuser_core.api.web.viewset import CustomPagination
from bkuser_core.audit.constants import OperationType
from bkuser_core.audit.utils import audit_general_log
from bkuser_core.bkiam.permissions import (
    IAMAction,
    IAMPermissionExtraInfo,
    ManageCategoryPermission,
    Permission,
    ViewCategoryPermission,
)
from bkuser_core.categories.constants import CategoryStatus, CategoryType, SyncTaskType
from bkuser_core.categories.exceptions import ExistsSyncingTaskError, FetchDataFromRemoteFailed
from bkuser_core.categories.loader import get_plugin_by_category
from bkuser_core.categories.models import ProfileCategory, SyncTask
from bkuser_core.categories.plugins.local.exceptions import DataFormatError
from bkuser_core.categories.signals import post_category_create, post_category_delete
from bkuser_core.categories.tasks import adapter_sync
from bkuser_core.common.error_codes import CoreAPIError, error_codes
from bkuser_core.departments.models import Department, DepartmentThroughModel
from bkuser_core.profiles.models import DynamicFieldInfo, Profile
from bkuser_core.user_settings.models import Setting
from bkuser_core.user_settings.signals import post_setting_create, post_setting_update

logger = logging.getLogger(__name__)


class CategoryMetasListApi(generics.ListAPIView):
    @classmethod
    def make_meta(cls, type_: CategoryType):
        return {
            "type": type_,
            "description": CategoryType.get_description(type_),
            "name": CategoryType.get_choice_label(type_),
        }

    def get(self, request, *args, **kwargs):
        """
        列表展示所有目录类型基本信息
        """
        metas = []
        for type_ in CategoryType.all():
            # 这里目前只返回创建目录类型的权限操作，后期应该可扩展
            try:
                action_id = IAMAction.get_action_by_category_type(type_)
            except KeyError:
                continue

            _meta = self.make_meta(type_)
            # Q：为什么这里需要手动判断权限，而不是通用 permission_classes？
            # A：因为这里的资源（目录类型）是没有对应实体，同时也没有在权限中心注册
            operator = get_operator(request)
            if not Permission().allow_action_without_resource(operator, action_id, raise_exception=False):
                _meta.update(
                    {
                        "authorized": False,
                        "extra_info": IAMPermissionExtraInfo.from_actions(
                            username=operator, action_ids=[action_id]
                        ).to_dict(),
                    }
                )
            metas.append(_meta)

        return Response(CategoryMetaOutputSLZ(metas, many=True).data)


class CategorySettingNamespaceListCreateUpdateApi(
    generics.ListAPIView, generics.UpdateAPIView, generics.CreateAPIView
):
    serializer_class = CategorySettingOutputSLZ
    permission_classes = [ManageCategoryPermission]

    def get_queryset(self):
        category_id = self.kwargs["id"]
        category = get_category(category_id)
        namespace = self.kwargs["namespace"]
        metas = list_setting_metas(category.type, None, namespace)
        return Setting.objects.filter(meta__in=metas, category_id=category_id)

    def post(self, request, *args, **kwargs):
        # 批量创建或更新
        slz = CategorySettingCreateInputSLZ(data=request.data, many=True)
        slz.is_valid(raise_exception=True)
        ns_settings = {d["key"]: d for d in slz.validated_data}

        # got metas
        category_id = self.kwargs["id"]
        category = get_category(category_id)
        namespace = self.kwargs["namespace"]
        metas = list_setting_metas(category.type, None, namespace)
        metas_dict = {m.key: m for m in metas}

        for key, setting in ns_settings.items():
            if key not in metas_dict:
                raise error_codes.CANNOT_FIND_SETTING_META.format(
                    f"can't find key={key} in namespace={namespace}", replace=True
                )
            meta = metas_dict[key]

            try:
                # 暂时忽略已创建报错
                setting, created = Setting.objects.update_or_create(
                    meta=meta, category=category, defaults={"value": setting["value"]}
                )
            except Exception:
                logger.exception(
                    "cannot create setting. [meta=%s, value=%s, category=%s]", meta, setting["value"], category
                )
                raise error_codes.CANNOT_CREATE_SETTING
            else:
                if created:
                    post_setting_create.send(
                        sender=self,
                        instance=setting,
                        operator=request.operator,
                        extra_values={"request": request},
                    )
                else:
                    # 仅当更新成功时才发送信号
                    post_setting_update.send(
                        sender=self,
                        instance=setting,
                        operator=request.operator,
                        extra_values={"request": request},
                    )

        settings = Setting.objects.filter(meta__in=metas, category_id=category_id).all()
        # TODO: 创建接口, 应该不需要返回. 确认前端是否有用到数据
        return Response(self.get_serializer_class()(settings, many=True).data)

    def put(self, request, *args, **kwargs):
        # 批量更新
        slz = CategoryNamespaceSettingUpdateInputSLZ(data=request.data, many=True)
        slz.is_valid(raise_exception=True)

        # [{key, value, enabled}]
        ns_settings = {d["key"]: d for d in slz.validated_data}

        category_id = self.kwargs["id"]
        category = get_category(category_id)
        namespace = self.kwargs["namespace"]
        metas = list_setting_metas(category.type, None, namespace)
        db_settings = Setting.objects.filter(meta__in=metas, category_id=category_id).all()

        # 更新已存在
        for s in db_settings:
            key = s.meta.key
            if key not in ns_settings:
                continue
            in_value = ns_settings[key]
            # 值变换才更新
            if s.value != in_value["value"] or s.enabled != in_value["enabled"]:
                s.value = in_value["value"]
                s.enabled = in_value["enabled"]
                s.save()
                post_setting_update.send(
                    sender=self,
                    instance=s,
                    operator=request.operator,
                    extra_values={"request": request},
                )
            del ns_settings[key]

        metas_map = {m.key: m for m in metas}
        # the left ns_settings are new settings
        for key, setting in ns_settings.items():
            if key not in metas_map:
                continue
            meta = metas_map[key]
            s, _ = Setting.objects.update_or_create(
                meta=meta,
                category_id=category_id,
                defaults={
                    "value": setting["value"],
                    "enabled": setting["enabled"],
                },
            )

            post_setting_create.send(
                sender=self,
                instance=s,
                operator=request.operator,
                extra_values={"request": request},
            )

        if ns_settings:
            # 有新建的setting内容，刷新db_settings
            db_settings = Setting.objects.filter(meta__in=metas, category_id=category_id).all()

        return Response(self.get_serializer_class()(db_settings, many=True).data)


class CategoryListCreateApi(generics.ListCreateAPIView):
    serializer_class = CategoryDetailOutputSLZ

    # TODO: 产品上应该返回全部列表, 展示这个人哪些有权限/哪些没权限
    # 而不是, 只返回有权限的
    def get_queryset(self):
        operator = get_operator(self.request)

        queryset = ProfileCategory.objects.filter(enabled=True)
        if settings.ENABLE_IAM:
            fs = Permission().make_category_filter(operator, IAMAction.VIEW_CATEGORY)
            queryset = queryset.filter(fs)

        return queryset

    def post(self, request, *args, **kwargs):
        """
        创建用户目录
        """
        slz = CategoryCreateInputSLZ(data=request.data)
        slz.is_valid(raise_exception=True)

        data = slz.validated_data

        if settings.ENABLE_IAM:
            # check permission
            operator = get_operator(request)
            action_id = IAMAction.get_action_by_category_type(data["type"])
            Permission().allow_action_without_resource(operator, action_id)

        instance = slz.save()

        # 默认添加到最后 TODO: 需要一个更优雅的实现
        max_order = ProfileCategory.objects.get_max_order()
        instance.order = max_order + 1
        instance.save(update_fields=["order"])
        post_category_create.send_robust(
            sender=self, instance=instance, operator=request.operator, extra_values={"request": request}
        )
        return Response(CategoryDetailOutputSLZ(instance).data, status=status.HTTP_201_CREATED)


class CategoryUpdateDeleteApi(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [ManageCategoryPermission]

    queryset = ProfileCategory.objects.all()
    lookup_url_kwarg = "id"

    @audit_general_log(operate_type=OperationType.UPDATE.value)
    def patch(self, request, *args, **kwargs):
        instance = self.get_object()
        slz = CategoryUpdateInputSLZ(instance, data=request.data, partial=True)
        slz.is_valid(raise_exception=True)
        self.perform_update(slz)

        return Response(CategoryDetailOutputSLZ(instance).data)

    def delete(self, request, *args, **kwargs):
        """删除用户目录"""
        instance = self.get_object()

        if instance.default:
            raise error_codes.CANNOT_DELETE_DEFAULT_CATEGORY
        # 不可删除非停用目录
        if instance.status != CategoryStatus.INACTIVE.value:
            raise error_codes.CANNOT_DELETE_ACTIVE_CATEGORY

        # 依赖 model 的 delete 方法, 执行软删除
        instance.delete()

        # 善后：回收站映射，软删除审计日志
        post_category_delete.send_robust(
            sender=self, instance=instance, operator=request.operator, extra_values={"request": request}
        )
        return Response(status=status.HTTP_200_OK)


class CategoryOperationTestConnectionApi(generics.CreateAPIView):
    permission_classes = [ManageCategoryPermission]

    queryset = ProfileCategory.objects.all()
    lookup_url_kwarg = "id"

    def post(self, request, *args, **kwargs):
        """测试连接"""
        slz = CategoryTestConnectionInputSLZ(data=request.data)
        slz.is_valid(raise_exception=True)

        instance = self.get_object()
        if instance.type not in [CategoryType.MAD.value, CategoryType.LDAP.value]:
            raise error_codes.TEST_CONNECTION_UNSUPPORTED

        # NOTE: 过于复杂, 在重构同步模块时, 更新这里的代码
        try:
            syncer_cls = get_plugin_by_category(instance).syncer_cls
        except Exception:
            logger.exception(
                "category<%s-%s-%s> load ldap client failed",
                instance.type,
                instance.display_name,
                instance.id,
            )
            raise error_codes.LOAD_LDAP_CLIENT_FAILED

        try:
            syncer_cls(instance.id, with_initialize_client=False).fetcher.client.initialize(**slz.validated_data)
        except Exception as e:
            logger.exception(
                "failed to test initialize category<%s-%s-%s>", instance.type, instance.display_name, instance.id
            )
            raise error_codes.TEST_CONNECTION_FAILED.format(str(e), replace=True)

        return Response()


class CategoryOperationTestFetchDataApi(generics.CreateAPIView):
    permission_classes = [ManageCategoryPermission]

    queryset = ProfileCategory.objects.all()
    lookup_url_kwarg = "id"

    def post(self, request, *args, **kwargs):
        """测试获取数据"""
        slz = CategoryTestFetchDataInputSLZ(data=request.data)
        slz.is_valid(raise_exception=True)

        instance = self.get_object()
        if instance.type not in [CategoryType.MAD.value, CategoryType.LDAP.value]:
            raise error_codes.TEST_CONNECTION_UNSUPPORTED

        try:
            syncer_cls = get_plugin_by_category(instance).syncer_cls
        except Exception:
            logger.exception(
                "category<%s-%s-%s> load data adapter failed",
                instance.type,
                instance.display_name,
                instance.id,
            )
            raise error_codes.LOAD_DATA_ADAPTER_FAILED

        try:
            syncer = syncer_cls(instance.id)
        except Exception as e:
            logger.exception(
                "failed to test initialize category<%s-%s-%s>", instance.type, instance.display_name, instance.id
            )
            raise error_codes.TEST_CONNECTION_FAILED.f(f"请确保连接设置正确 {str(e)}")

        try:
            syncer.fetcher.test_fetch_data(slz.validated_data)
        except Exception as e:  # pylint: disable=broad-except
            logger.exception(
                "failed to fetch data from category<%s-%s-%s>", instance.type, instance.display_name, instance.id
            )
            error_detail = f" ({type(e).__module__}.{type(e).__name__}: {str(e)})"
            raise error_codes.TEST_FETCH_DATA_FAILED.f(error_detail)

        return Response()


class CategoryOperationExportTemplateApi(generics.RetrieveAPIView):
    permission_classes = [ManageCategoryPermission]

    def get(self, request, *args, **kwargs):
        """生成excel导入模板样例文件"""
        fields = DynamicFieldInfo.objects.filter(enabled=True).all()
        data = FieldOutputSLZ(fields, many=True).data
        exporter = ProfileExcelExporter(
            load_workbook(settings.EXPORT_ORG_TEMPLATE), settings.EXPORT_EXCEL_FILENAME + "_org_tmpl", data
        )
        exporter.update_sheet_titles(exclude_keys=["last_login_time", "create_time"])

        return exporter.to_response()


class CategoryOperationExportApi(generics.RetrieveAPIView):
    permission_classes = [ManageCategoryPermission]

    queryset = ProfileCategory.objects.all()
    lookup_url_kwarg = "id"

    def get(self, request, *args, **kwargs):
        """导出组织架构"""
        slz = CategoryExportInputSLZ(data=self.request.query_params)
        slz.is_valid(raise_exception=True)

        data = slz.validated_data
        department_ids = data["department_ids"]

        category = self.get_object()
        if not category.type == CategoryType.LOCAL.value:
            raise error_codes.LOCAL_CATEGORY_NEEDS_EXCEL_FILE

        ids = Department.tree_objects.get_queryset_descendants(
            queryset=Department.objects.filter(id__in=department_ids), include_self=True
        ).values_list("id", flat=True)

        profile_ids = DepartmentThroughModel.objects.filter(department_id__in=ids).values_list("profile_id", flat=True)
        profiles = Profile.objects.filter(id__in=profile_ids)

        # FIXME: profile slz should contains?
        all_profiles = CategoryExportProfileOutputSLZ(profiles, many=True).data

        fields = DynamicFieldInfo.objects.filter(enabled=True).all()
        fields_data = FieldOutputSLZ(fields, many=True).data
        exporter = ProfileExcelExporter(
            load_workbook(settings.EXPORT_ORG_TEMPLATE), settings.EXPORT_EXCEL_FILENAME + "_org", fields_data
        )
        exporter.update_profiles(all_profiles)

        return exporter.to_response()


class CategoryOperationSyncOrImportApi(generics.CreateAPIView):
    permission_classes = [ManageCategoryPermission]
    queryset = ProfileCategory.objects.filter(enabled=True)
    lookup_url_kwarg = "id"

    parser_classes: List = [
        FileUploadParser,
    ]

    # @audit_general_log(operate_type=OperationType.IMPORT.value)
    @audit_general_log(operate_type=OperationType.SYNC.value)
    def post(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.type == CategoryType.LOCAL.value:
            return self._local_category_do_import(request, instance)

        return self._not_local_category_do_sync(request, instance)

    def _local_category_do_import(self, request, instance):
        """向本地目录导入数据文件"""
        query_slz = CategoryFileImportQuerySLZ(data=request.query_params)
        query_slz.is_valid(raise_exception=True)

        slz = CategoryFileImportInputSLZ(data=request.data)
        slz.is_valid(raise_exception=True)

        try:
            task_id = SyncTask.objects.register_task(
                category=instance, operator=request.operator, type_=SyncTaskType.MANUAL
            ).id
        except ExistsSyncingTaskError as e:
            raise error_codes.CREATE_SYNC_TASK_FAILED.f(str(e))

        instance_id = instance.id
        params = {
            "raw_data_file": slz.validated_data["file"],
            "is_overwrite": query_slz.validated_data["is_overwrite"],
            "language": translation.get_language(),
        }
        try:
            # TODO: FileField 可能不能反序列化, 所以可能不能传到 celery 执行
            adapter_sync(instance_id, operator=request.operator, task_id=task_id, **params)
        except DataFormatError as e:
            logger.exception(
                "failed to sync data, data format error. [instance_id=%s, operator=%s, task_id=%s, params=%s]",
                instance_id,
                request.operator,
                task_id,
                params,
            )
            raise error_codes.SYNC_DATA_FAILED.format(str(e), replace=True)
        except Exception as e:
            logger.exception(
                "failed to sync data. [instance_id=%s, operator=%s, task_id=%s, params=%s]",
                instance_id,
                request.operator,
                task_id,
                params,
            )
            raise error_codes.SYNC_DATA_FAILED.format(str(e), replace=True)

        return Response(CategorySyncResponseOutputSLZ({"task_id": task_id}).data)

    def _not_local_category_do_sync(self, request, instance):
        """同步目录"""
        try:
            task_id = SyncTask.objects.register_task(
                category=instance, operator=request.operator, type_=SyncTaskType.MANUAL
            ).id
        except ExistsSyncingTaskError as e:
            logger.exception(
                "failed to register sync task. [instance.id=%s], operator=%s", instance.id, request.operator
            )
            raise error_codes.LOAD_DATA_FAILED.f(str(e))

        try:
            adapter_sync.apply_async(
                kwargs={"instance_id": instance.id, "operator": request.operator, "task_id": task_id}
            )
        except FetchDataFromRemoteFailed as e:
            logger.exception(
                "failed to sync data. fetch data from remote fail. [instance.id=%s, operator=%s, task_id=%s]",
                instance.id,
                request.operator,
                task_id,
            )
            error_detail = f" ({type(e).__module__}.{type(e).__name__}: {str(e)})"
            raise error_codes.SYNC_DATA_FAILED.f(error_detail)
        except CoreAPIError:
            raise
        except Exception as e:
            logger.exception(
                "failed to sync data. [instance.id=%s, operator=%s, task_id=%s]",
                instance.id,
                request.operator,
                task_id,
            )
            raise error_codes.SYNC_DATA_FAILED.f(f"{e}")

        return Response(CategorySyncResponseOutputSLZ({"task_id": task_id}).data)


class CategoryOperationSwitchOrderApi(generics.UpdateAPIView):
    permission_classes = [ManageCategoryPermission]
    queryset = ProfileCategory.objects.filter(enabled=True)
    lookup_url_kwarg = "id"

    @audit_general_log(operate_type=OperationType.UPDATE.value)
    def patch(self, request, *args, **kwargs):
        instance = self.get_object()

        another_category = ProfileCategory.objects.get(id=kwargs["another_id"])

        # switch
        instance.order, another_category.order = another_category.order, instance.order
        instance.save(update_fields=["order"])
        another_category.save(update_fields=["order"])

        return Response()


class CategoryProfileListApi(generics.ListAPIView):
    permission_classes = [ViewCategoryPermission]
    pagination_class = CustomPagination
    serializer_class = ProfileDetailOutputSLZ

    def get_queryset(self):
        slz = CategoryProfileListInputSLZ(data=self.request.query_params)
        slz.is_valid(raise_exception=True)
        data = slz.validated_data

        # filter by category_id
        category_id = self.kwargs["id"]
        queryset = Profile.objects.filter(category_id=category_id, enabled=True)

        # filter by keyword
        keyword = data.get("keyword")
        if keyword:
            # NOTE: 这里相对原来的差异, 抹掉了 id__icontains 的搜索
            queryset = queryset.filter(Q(username__icontains=keyword) | Q(display_name__icontains=keyword))

        # 首页展示不关联部门的用户列表
        has_no_department = data.get("has_no_department", False)
        if has_no_department:
            queryset = queryset.filter(departments__isnull=True)
            # SQL:
            # SELECT COUNT(*) AS `__count` FROM `profiles_profile`
            # LEFT OUTER JOIN `departments_department_profiles`
            # ON (`profiles_profile`.`id` = `departments_department_profiles`.`profile_id`)
            # WHERE (`profiles_profile`.`category_id` = 1
            #        AND `profiles_profile`.`enabled`
            #        AND `departments_department_profiles`.`department_id` IS NULL);

        # do prefetch
        queryset = queryset.prefetch_related("departments", "leader")
        return queryset


class CategoryDepartmentListApi(generics.ListAPIView):

    permission_classes = [ManageCategoryPermission]
    # NOTE: 这里跟原先的区别, 全部返回的 with_ancestors=true
    serializer_class = DepartmentsWithChildrenAndAncestorsOutputSLZ
    pagination_class = CustomPagination

    def get_queryset(self):
        category_id = self.kwargs["id"]
        queryset = Department.objects.filter(category_id=category_id)

        slz = CategoryDepartmentListInputSLZ(data=self.request.query_params)
        slz.is_valid(raise_exception=True)
        data = slz.validated_data

        keyword = data.get("keyword")
        if keyword:
            if keyword.isdigit():
                queryset = queryset.filter(Q(name__icontains=keyword) | Q(id=keyword))
            else:
                queryset = queryset.filter(name__icontains=keyword)

        return queryset

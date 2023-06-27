# -*- coding: utf-8 -*-
"""
Tencent is pleased to support the open source community by making 蓝鲸智云PaaS平台社区版 (BlueKing PaaS Community
Edition) available.
Copyright (C) 2017-2020 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the MIT License (the "License"); you may not use this file except in compliance with the License.
You may obtain a copy of the License at
http://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
specific language governing permissions and limitations under the License.
"""
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import TYPE_CHECKING, List

from django.http import HttpResponse
from django.utils.translation import ugettext_lazy as _
from openpyxl.styles import Alignment, Font, colors
from openpyxl.styles.numbers import FORMAT_TEXT
from rest_framework import serializers

from bkuser_core.profiles.constants import DynamicFieldTypeEnum

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)


def get_options_values_by_key(options: list, keys: list):
    if not options:
        return None

    values = []
    for k in keys:
        for pair in options:
            if pair["id"] == k:
                values.append(pair["value"])

    return values


class LeaderSerializer(serializers.Serializer):
    display_name = serializers.CharField()
    id = serializers.IntegerField()
    username = serializers.CharField()


class SubDepartmentSerializer(serializers.Serializer):
    id = serializers.IntegerField(required=False)
    name = serializers.CharField(required=False)
    order = serializers.IntegerField(required=False)
    full_name = serializers.CharField(required=False)
    has_children = serializers.BooleanField(required=False)


class ProfileExportSerializer(serializers.Serializer):
    display_name = serializers.CharField()
    username = serializers.CharField()
    leader = LeaderSerializer(many=True)
    department_name = SubDepartmentSerializer(many=True, source="departments")
    staff_status = serializers.CharField(required=False)
    status = serializers.CharField(required=False)
    extras = serializers.JSONField(default={})
    telephone = serializers.CharField()
    email = serializers.CharField()
    qq = serializers.CharField(required=False)
    position = serializers.IntegerField(required=False)
    wx_userid = serializers.CharField()
    iso_code = serializers.CharField()
    country_code = serializers.CharField()
    last_login_time = serializers.CharField()
    create_time = serializers.DateTimeField()
    account_expiration_date = serializers.DateField()

    def to_representation(self, instance):
        data = super().to_representation(instance)

        data["leader"] = ",".join(x["username"] for x in data["leader"])
        data["department_name"] = ",".join([x["full_name"] for x in data["department_name"]])
        return data


@dataclass
class ProfileExcelExporter:
    """导出"""

    workbook: "Workbook"
    exported_file_name: str
    fields: list
    title_row_index: int = 2

    def __post_init__(self):
        self.first_sheet = self.workbook.worksheets[0]
        # 样式加载
        self.first_sheet.alignment = Alignment(wrapText=True)
        # 初始化全表的单元格数据格式
        # 将单元格设置为纯文本模式，预防DDE
        for columns in self.first_sheet.columns:
            for cell in columns:
                cell.number_format = FORMAT_TEXT

    def update_profiles(self, profiles: List[dict], extra_infos: dict = None):
        field_col_map = self.update_sheet_titles()

        for p_index, p in enumerate(profiles):

            exported_profile = ProfileExportSerializer(p).data
            for f_index, f in enumerate(self.fields):
                # field_name = f["name"]
                field_name = f["key"]

                try:
                    if f["builtin"]:
                        raw_value = exported_profile[field_name]
                    else:
                        raw_value = exported_profile["extras"][field_name]
                except KeyError:
                    # 当无法从当前用户属性中找到对应字段时，尝试从 extra_infos 中获取
                    if extra_infos is None:
                        logger.exception("failed to get value from extra_infos field<%s>, extra_infos is empty", f)
                        continue

                    try:
                        extra = extra_infos.get(p["id"]) or extra_infos.get(str(p["id"])) or {}
                        raw_value = extra[field_name]
                        if isinstance(raw_value, datetime):
                            raw_value = raw_value.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
                    except KeyError:
                        logger.exception("failed to get value from extra_infos field<%s>, key missing", f)
                        continue

                value = raw_value
                # options 存储值为 key， 但是 Excel 交互值为 value
                if f["type"] == DynamicFieldTypeEnum.ONE_ENUM.value:
                    value = ",".join(get_options_values_by_key(f["options"], [raw_value]))
                elif f["type"] == DynamicFieldTypeEnum.MULTI_ENUM.value:
                    value = ",".join(get_options_values_by_key(f["options"], raw_value))

                # 为电话添加国际号码段
                if f["name"] == "telephone":
                    value = f'+{exported_profile["country_code"]}{exported_profile[field_name]}'

                if raw_value is None:
                    continue

                self.first_sheet.cell(
                    row=p_index + self.title_row_index + 1, column=field_col_map[f["display_name"]], value=value
                )

    def update_sheet_titles(self, exclude_keys: List[str] = []):
        """更新表格标题"""
        required_fields = [x for x in self.fields if x["builtin"] and x["name"] not in exclude_keys]
        not_required_field_names = [
            x["display_name"] for x in self.fields if not x["builtin"] and x["name"] not in exclude_keys
        ]

        field_col_map = {}

        # NOTE: 这里required_field_names 跟excel表格中对不上了, 两边的意义不一样(模型定义是required, 但在excel导入是非必填的)
        # 所以, 我们这里使用枚举的方式, 保证导入/导出的标红title一致

        red_ft = Font(color=colors.COLOR_INDEX[2])
        black_ft = Font(color=colors.BLACK)
        for index, field in enumerate(required_fields):
            field_name = field["display_name"]
            field_key = field["key"]
            column = index + 1
            _cell = self.first_sheet.cell(
                row=self.title_row_index,
                column=column,
                value=field_name,
            )
            if field_key in ("display_name", "username", "email", "telephone", "department_name"):
                _cell.font = red_ft
            else:
                _cell.font = black_ft

            field_col_map[field_name] = index + 1

        for index, field_name in enumerate(not_required_field_names):
            column = index + 1 + len(required_fields)
            _cell = self.first_sheet.cell(
                row=self.title_row_index,
                column=column,
                value=field_name,
            )
            _cell.font = black_ft
            field_col_map[field_name] = column

        return field_col_map

    def to_response(self) -> HttpResponse:
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f"attachment;filename={self.exported_file_name}.xlsx"
        self.workbook.save(response)
        return response


class LoginLogExcelExporter:
    """登录审计日志导出"""

    workbook: "Workbook"
    exported_file_name: str
    fields: list = [
        {
            "title": _("登录用户"),
            "name": "username",
        },
        {
            "title": _("用户全名"),
            "name": "display_name",
        },
        {
            "title": _("登录时间"),
            "name": "datetime",
        },
        {
            "title": _("登录来源IP"),
            "name": "ip",
        },
        {
            "title": _("登录状态"),
            "name": "status",
        },
        {
            "title": _("登录失败原因"),
            "name": "reason",
        },
    ]
    title_row_index: int = 1

    def __init__(self, webhook: "Workbook", exported_file_name: str):
        self.workbook = webhook
        self.exported_file_name = exported_file_name

        self.first_sheet = self.workbook.worksheets[0]
        # 样式加载
        self.first_sheet.alignment = Alignment(wrapText=True)
        # 初始化全表的单元格数据格式
        # 将单元格设置为纯文本模式，预防DDE
        for columns in self.first_sheet.columns:
            for cell in columns:
                cell.number_format = FORMAT_TEXT

    def to_response(self) -> HttpResponse:
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f"attachment;filename={self.exported_file_name}.xlsx"
        self.workbook.save(response)
        return response

    def add_records(self, records):
        self._update_sheet_titles()
        for p_index, p in enumerate(records):
            for f_index, field in enumerate(self.fields):
                field_name = field["name"]
                value = p[field_name]
                self.first_sheet.cell(row=p_index + self.title_row_index + 1, column=f_index + 1, value=value)

    def _update_sheet_titles(self):
        """更新表格标题"""
        black_ft = Font(color=colors.BLACK)

        for index, field in enumerate(self.fields):
            # column = index + 1 + len(self.fields)
            column = index + 1
            _cell = self.first_sheet.cell(
                row=self.title_row_index,
                column=column,
                value=str(field["title"]),
            )
            _cell.font = black_ft

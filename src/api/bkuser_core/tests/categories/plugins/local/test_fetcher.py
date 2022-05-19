# -*- coding: utf-8 -*-
"""
TencentBlueKing is pleased to support the open source community by making 蓝鲸智云-用户管理(Bk-User) available.
Copyright (C) 2017-2021 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the MIT License (the "License"); you may not use this file except in compliance with the License.
You may obtain a copy of the License at http://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
specific language governing permissions and limitations under the License.
"""
from pathlib import Path
from shutil import copyfile

import pytest
from openpyxl import load_workbook

from .constants import COMMON_TITLES
from bkuser_core.categories.plugins.local.syncer import ExcelFetcher

pytestmark = pytest.mark.django_db


class TestExcelFetcher:
    test_tmpl_file_path = Path(__file__).parents[0] / "assets/empty.xlsx"
    test_tmp_file_path = Path(__file__).parents[0] / "assets/temp.xlsx"

    @pytest.fixture
    def fetcher(self, local_category):
        return ExcelFetcher(category_id=local_category.id)

    @pytest.fixture
    def fake_excel(self, raw_data: list):
        copyfile(self.test_tmpl_file_path, self.test_tmp_file_path)

        excel_file = load_workbook(self.test_tmp_file_path)
        empty_sheet = excel_file.worksheets[0]
        for line, raw_data_inline in enumerate(raw_data):
            for index, raw_data_in_cell in enumerate(raw_data_inline):
                empty_sheet.cell(
                    row=line + 2,
                    column=index + 1,
                    value=raw_data_in_cell,
                )

        excel_file.save(self.test_tmp_file_path)
        yield empty_sheet

        Path(self.test_tmp_file_path).unlink()

    @pytest.mark.parametrize(
        "raw_data,expected",
        [
            (
                [COMMON_TITLES, [*["aaa"] * 6, "aaa/bbb"], [*["bbb"] * 6, "aaa/bbb"]],
                (
                    [(*["aaa"] * 6, "aaa/bbb"), (*["bbb"] * 6, "aaa/bbb")],
                    [["aaa", "bbb"]],
                ),
            ),
            (
                [COMMON_TITLES, [*["aaa"] * 6, "aaa/bbb"], [*["bbb"] * 6, "aaa/ccc"]],
                (
                    [(*["aaa"] * 6, "aaa/bbb"), (*["bbb"] * 6, "aaa/ccc")],
                    [["aaa", "bbb"], ["aaa", "ccc"]],
                ),
            ),
        ],
    )
    def test_fetch(self, fetcher, raw_data, expected, fake_excel):
        """test fetch data"""
        assert fetcher.fetch(self.test_tmp_file_path) == expected

    @pytest.mark.parametrize(
        "raw_data",
        [
            [["未知表头"], [*["aaa"] * 6, "aaa/bbb"], [*["bbb"] * 6, "aaa/bbb"]],
            [
                ["用户名", "全名", "未知表头"],
                [*["aaa"] * 6, "aaa/bbb"],
                [*["bbb"] * 6, "aaa/bbb"],
            ],
        ],
    )
    def test_unknown_key_fetch(self, fetcher, raw_data, fake_excel):
        """test fetch data"""
        with pytest.raises(ValueError):
            fetcher.fetch(self.test_tmp_file_path)

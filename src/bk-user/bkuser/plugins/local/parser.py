# -*- coding: utf-8 -*-
"""
TencentBlueKing is pleased to support the open source community by making 蓝鲸智云-用户管理(Bk-User) available.
Copyright (C) 2017-2021 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the MIT License (the "License"); you may not use this file except in compliance with the License.
You may obtain a copy of the License at http://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
specific language governing permissions and limitations under the License.
"""
from collections import Counter
from typing import List

import phonenumbers
from django.conf import settings
from django.utils.translation import gettext_lazy as _
from openpyxl.workbook import Workbook

from bkuser.plugins.base import PluginLogger
from bkuser.plugins.local.constants import USERNAME_REGEX
from bkuser.plugins.local.exceptions import (
    CustomColumnNameInvalid,
    DuplicateColumnName,
    DuplicateUsername,
    InvalidLeader,
    InvalidOrganization,
    InvalidUsername,
    RequiredFieldIsEmpty,
    SheetColumnsNotMatch,
    UserSheetNotExists,
)
from bkuser.plugins.local.utils import gen_dept_code
from bkuser.plugins.models import RawDataSourceDepartment, RawDataSourceUser


class LocalDataSourceDataParser:
    """本地数据源数据解析器"""

    # 用户表名称
    user_sheet_name = "users"
    # 第一行是填写必读，第二行才是列名（1-based）
    col_name_row_idx = 2
    # 第三行开始，才是用户数据（1-based)
    user_data_min_row_idx = 3
    # 用户名索引（0-based）
    username_idx_in_row = 0
    # 组织列索引（0-based）
    organization_idx_in_row = 4

    # 内建字段列名
    builtin_col_names = [
        "用户名/username",
        "姓名/full_name",
        "邮箱/email",
        "手机号/phone_number",
        "组织/organizations",
        "直接上级/leaders",
    ]
    # 内建字段列长度
    builtin_col_length = len(builtin_col_names)

    # NOTE 下列字段在加载到 workbook 后填充
    # 自定义字段列名
    custom_col_names: List[str] = []
    # 完整的字段列名 = 内建字段列名 + 自定义字段列名
    all_col_names: List[str] = []
    # 完整的字段名称
    all_field_names: List[str] = []
    # 有效字段列长度（内建字段 + 自定义字段）
    valid_col_length = 0

    # 必填字段列名，自定义必填字段不在解析器中校验
    required_field_names = [
        "username",
        "full_name",
        "email",
        "phone_number",
    ]

    def __init__(self, logger: PluginLogger, workbook: Workbook):
        self.logger = logger
        self.workbook = workbook
        self.departments: List[RawDataSourceDepartment] = []
        self.users: List[RawDataSourceUser] = []
        self.is_parsed = False

    def parse(self):
        """预解析部门 & 用户数据"""
        self._validate_and_prepare()
        self._parse_departments()
        self._parse_users()
        self.is_parsed = True

    def get_departments(self) -> List[RawDataSourceDepartment]:
        return self.departments

    def get_users(self) -> List[RawDataSourceUser]:
        return self.users

    def _validate_and_prepare(self):
        """检查表格格式，确保后续可正常解析"""
        self._validate_sheet()
        self._validate_columns()
        self._validate_users()

    def _validate_sheet(self):
        # 确保用户表确实存在
        if self.user_sheet_name not in self.workbook.sheetnames:
            raise UserSheetNotExists(_("待导入文件中不存在用户表"))

        self.sheet = self.workbook[self.user_sheet_name]

    def _validate_columns(self):
        # 1. 检查表头是否正确
        sheet_col_names = [cell.value for cell in self.sheet[self.col_name_row_idx]]
        # 前 N 个是内建字段，必须存在
        builtin_col_length = len(self.builtin_col_names)
        if sheet_col_names[:builtin_col_length] != self.builtin_col_names:
            raise SheetColumnsNotMatch(_("待导入文件中用户表格式异常"))

        # N 个之后，是可能存在的自定义字段
        self.custom_col_names = sheet_col_names[builtin_col_length:]
        self.all_col_names = self.builtin_col_names + self.custom_col_names
        self.valid_col_length = len(self.all_col_names)
        self.logger.info(f"all column names parsed from workbook: {self.all_col_names}")  # noqa: G004

        # 2. 检查自定义字段是否符合格式，格式：display_name/field_name
        for col_name in self.custom_col_names:
            if not col_name:
                raise CustomColumnNameInvalid(_("存在值为空的自定义字段列名，请下载使用最新的模板").format(col_name))

            display_name, __, field_name = col_name.partition("/")
            if not (display_name and field_name):
                raise CustomColumnNameInvalid(_("自定义字段 {} 格式不合法，参考格式：年龄/age").format(col_name))

        # 获取所有的字段名
        self.all_field_names = [n.split("/")[-1] for n in self.all_col_names]
        self.logger.info(f"all field names parsed from workbook: {self.all_field_names}")  # noqa: G004

        # 3. 检查是否有重复列
        if duplicate_col_names := [n for n, cnt in Counter(sheet_col_names).items() if cnt > 1]:
            raise DuplicateColumnName(_("待导入文件中存在重复列名：{}").format(", ".join(duplicate_col_names)))

    def _validate_users(self):
        all_usernames = []
        for idx, cell_values in enumerate(
            self.sheet.iter_rows(min_row=self.user_data_min_row_idx, max_col=self.valid_col_length, values_only=True),
            start=self.user_data_min_row_idx,
        ):
            if not any(cell_values):
                self.logger.warning(f"empty row found at line {idx} in sheet, skip...")  # noqa: G004
                continue

            info = dict(zip(self.all_field_names, cell_values, strict=True))
            # 1. 检查所有必填字段是否有值（注：自定义字段必填在后续的流程中检查）
            for field_name in self.required_field_names:
                if not info.get(field_name):
                    raise RequiredFieldIsEmpty(_("待导入文件中必填字段 {} 存在空值").format(field_name))

            username = info["username"]
            # 2. 检查用户名是否合法
            if not USERNAME_REGEX.fullmatch(username):
                raise InvalidUsername(
                    _(
                        "用户名 {} 不符合命名规范: 由3-32位字母、数字、下划线(_)、点(.)、连接符(-)字符组成，以字母或数字开头及结尾",  # noqa: E501
                    ).format(username)
                )

            # 3. 检查用户不能是自己的 leader
            if (leaders := info.get("leaders")) and username in [ld.strip() for ld in leaders.split(",")]:
                raise InvalidLeader(_("待导入文件中用户 {} 不能是自己的直接上级").format(username))

            all_usernames.append(username.lower())

        # 4. 检查用户名是否有重复的（以大小写不敏感的方式检查）
        if duplicate_usernames := [n for n, cnt in Counter(all_usernames).items() if cnt > 1]:
            raise DuplicateUsername(
                _(
                    "待导入文件中存在重复用户名：{}（该检查大小写不敏感）",
                ).format(", ".join(duplicate_usernames))
            )

    def _parse_departments(self):
        organizations = set()
        for cell_values in self.sheet.iter_rows(
            min_row=self.user_data_min_row_idx, max_col=self.valid_col_length, values_only=True
        ):
            if not any(cell_values):
                continue

            username, user_orgs = cell_values[self.username_idx_in_row], cell_values[self.organization_idx_in_row]
            if not user_orgs:
                self.logger.info(f"username {username} not provide organization, skip...")  # noqa: G004
                continue

            for org in user_orgs.split(","):
                cur_org = org.strip()
                if not all(cur_org.split("/")):
                    raise InvalidOrganization(
                        _(
                            "用户 {} 组织路径 {} 不合法：不得以 / 开头或结尾或存在连续的 / 字符",
                        ).format(username, cur_org)
                    )

                organizations.add(cur_org)
                # 所有的父部门都要被添加进来
                while "/" in cur_org:
                    cur_org, __, __ = cur_org.rpartition("/")
                    organizations.add(cur_org.strip())

        # 组织路径：本数据源部门 Code 映射表
        org_code_map = {org: gen_dept_code(org) for org in organizations}
        for org in organizations:
            parent_org, __, dept_name = org.rpartition("/")
            self.departments.append(
                RawDataSourceDepartment(
                    code=org_code_map[org],
                    name=dept_name,
                    parent=org_code_map.get(parent_org),
                    # 注：本地数据源导入，不支持设置部门的 extras 信息
                    extras={},
                )
            )

    def _parse_users(self):
        for cell_values in self.sheet.iter_rows(
            min_row=self.user_data_min_row_idx, max_col=self.valid_col_length, values_only=True
        ):
            if not any(cell_values):
                continue

            properties = dict(zip(self.all_field_names, cell_values, strict=True))

            departments, leaders = [], []
            if organizations := properties.pop("organizations"):
                for org in organizations.split(","):
                    if org := org.strip():
                        departments.append(gen_dept_code(org))

            if leader_names := properties.pop("leaders"):
                for ld in leader_names.split(","):
                    if ld := ld.strip():
                        # xlsx 中填写的是 leader 的 username，但在本地数据源中，username 就是 code
                        leaders.append(ld)

            phone_number = str(properties.pop("phone_number"))
            # 默认认为是不带国际代码的
            phone, country_code = phone_number, settings.DEFAULT_PHONE_COUNTRY_CODE
            if phone_number.startswith("+"):
                ret = phonenumbers.parse(phone_number)
                phone, country_code = str(ret.national_number), str(ret.country_code)

            properties.update({"phone": phone, "phone_country_code": country_code})

            # 格式化，将所有非 None 字段都转成 str 类型
            properties = {k: str(v) for k, v in properties.items() if v is not None}
            self.users.append(
                RawDataSourceUser(
                    # 本地数据源用户，code 就是 username
                    code=properties["username"],
                    properties=properties,
                    leaders=leaders,
                    departments=departments,
                )
            )

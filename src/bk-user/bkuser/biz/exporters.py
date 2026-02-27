# -*- coding: utf-8 -*-
# TencentBlueKing is pleased to support the open source community by making
# 蓝鲸智云 - 用户管理 (bk-user) available.
# Copyright (C) 2017 Tencent. All rights reserved.
# Licensed under the MIT License (the "License"); you may not use this file except
# in compliance with the License. You may obtain a copy of the License at
#
#     http://opensource.org/licenses/MIT
#
# Unless required by applicable law or agreed to in writing, software distributed under
# the License is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND,
# either express or implied. See the License for the specific language governing permissions and
# limitations under the License.
#
# We undertake not to change the open source license (MIT license) applicable
# to the current version of the project delivered to anyone in the future.
from itertools import groupby
from typing import Any, Dict, List

from django.conf import settings
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font, colors
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from bkuser.apps.data_source.models import (
    DataSource,
    DataSourceDepartment,
    DataSourceDepartmentRelation,
    DataSourceDepartmentUserRelation,
    DataSourceUser,
    DataSourceUserLeaderRelation,
)
from bkuser.apps.tenant.constants import UserFieldDataType
from bkuser.apps.tenant.models import TenantUserCustomField


class UserExcelWriter:
    """负责 Excel 文件的结构定义，样式渲染和数据写入"""

    # 模版中字段名行索引
    HEADER_ROW_IDX = 2
    # 提示信息行索引
    TIP_ROW_IDX = 1
    # 列的默认宽度
    DEFAULT_COL_WIDTH = 40

    def __init__(self, custom_fields: List[TenantUserCustomField]):
        self.custom_fields = custom_fields
        self.field_option_map = self._build_option_map(custom_fields)

        self.workbook = load_workbook(settings.EXPORT_ORG_TEMPLATE)
        self.sheet: Worksheet = self.workbook["users"]
        self.sheet.alignment = Alignment(wrapText=True)

        # 基础列数（用户名, 姓名, 邮箱, 手机, 组织, 上级）
        self.base_col_count = 6
        self._init_sheet_structure()

    def add_row(self, base_info: List[str], extras: Dict[str, Any]):
        # Q: 为什么不能这样写 extras.get(field.name) or ""
        # A: 这会导致部分类型的零值被转为空字符串，不符合预期，比如 类型是NUMBER, 0 应该是 "0"，而非 ""
        extra_values = [
            self._transform_custom_field_value(field, extras.get(field.name, "")) for field in self.custom_fields
        ]
        self.sheet.append(base_info + extra_values)  # noqa: PERF401 sheet isn't a list

    def get_workbook(self) -> Workbook:
        return self.workbook

    def _init_sheet_structure(self):
        # openpyxl 列索引从 1 开始
        for col_idx, field in enumerate(self.custom_fields, start=self.base_col_count + 1):
            col_letter = get_column_letter(col_idx)

            # 设置字段名
            name_cell = self.sheet.cell(row=self.HEADER_ROW_IDX, column=col_idx)
            name_cell.value = f"{field.display_name}/{field.name}"
            # 设置为垂直居中 + 自动换行
            name_cell.alignment = Alignment(vertical="center", wrapText=True)
            # 如果是必填列，列名设置为红色
            if field.required:
                name_cell.font = Font(color=colors.COLOR_INDEX[2])

            # 设置提示信息(字段名的上一行）
            tip_cell = self.sheet.cell(row=self.TIP_ROW_IDX, column=col_idx)
            tip_cell.value = self._get_field_tip(field)
            # 设置吸底 + 自动换行
            tip_cell.alignment = Alignment(vertical="bottom", wrapText=True)

            # 设置默认列宽
            self.sheet.column_dimensions[col_letter].width = self.DEFAULT_COL_WIDTH

        self._set_all_columns_to_text_format()

    def _transform_custom_field_value(
        self, field: TenantUserCustomField, value: List[str | int | float] | str | int | float
    ) -> str:
        """
        转换自定义字段的值，以字符串输出；注意枚举做 id 与 value 的映射输出处理
        """
        # 对于单枚举（""）、多枚举（[]）、字符串("") 类型，当其为零值时，可提前返回空字符串
        # 对于数值（0）类型，则需要按照正常处理，即 str(0)，不能返回空字符串
        # Note: 无值，value 本身就是空字符串
        if field.data_type != UserFieldDataType.NUMBER and not value:
            return ""

        # 单枚举，则选项映射
        if field.data_type == UserFieldDataType.ENUM:
            return self.field_option_map[field.name][value]  # type: ignore

        # 多枚举，则选项映射并逗号拼接
        if field.data_type == UserFieldDataType.MULTI_ENUM:
            return ",".join([self.field_option_map[field.name][opt_id] for opt_id in value])  # type: ignore

        # 非枚举类型，直接转换为字符串
        return str(value)

    def _set_all_columns_to_text_format(self):
        # 将单元格设置为纯文本模式，防止出现类型转换
        # ref: https://stackoverflow.com/questions/57492559
        for idx, _ in enumerate(self.sheet.columns):
            self.sheet.column_dimensions[get_column_letter(idx + 1)].number_format = FORMAT_TEXT

    @staticmethod
    def _get_field_tip(field: TenantUserCustomField) -> str:
        opts = ", ".join(opt["value"] for opt in field.options)
        if field.data_type == UserFieldDataType.ENUM:
            return f"单选枚举（One of {opts}）"
        if field.data_type == UserFieldDataType.MULTI_ENUM:
            return f"多选枚举，多个值以英文逗号分隔（Any of {opts}, separated by commas）"
        if field.data_type == UserFieldDataType.NUMBER:
            return "数据类型：数值（DataType: Number）"
        if field.data_type == UserFieldDataType.STRING:
            return "数据类型：字符串（DataType: String）"
        return ""

    @staticmethod
    def _build_option_map(fields: List[TenantUserCustomField]) -> Dict[str, Dict[str, str]]:
        return {f.name: {opt["id"]: opt["value"] for opt in f.options} for f in fields}


class DataSourceUserExporter:
    """导出数据源用户 & 组织信息"""

    def __init__(self, data_source: DataSource):
        self.data_source = data_source
        self.users = DataSourceUser.objects.filter(data_source=data_source)
        self.custom_fields = TenantUserCustomField.objects.filter(tenant_id=data_source.owner_tenant_id)
        self.writer = UserExcelWriter(self.custom_fields)

    @staticmethod
    def get_template(tenant_id: str) -> Workbook:
        custom_fields = list(TenantUserCustomField.objects.filter(tenant_id=tenant_id))
        writer = UserExcelWriter(custom_fields)

        # 填充自定义字段默认值以供参考
        extras = {f.name: f.default for f in custom_fields}
        base_info = [
            "zhangsan",
            "张三",
            "zhangsan@qq.com",
            "+8613512345678",
            "公司/部门A,公司/部门B",
            "lisi,wangwu",
        ]

        writer.add_row(base_info, extras)
        return writer.get_workbook()

    def export(self) -> Workbook:
        dept_org_map = self._build_dept_org_map()
        user_departments_map = self._build_user_departments_map()
        user_leaders_map = self._build_user_leaders_map()
        user_username_map = self._build_user_username_map()

        for u in self.users:
            phone = f"+{u.phone_country_code}{u.phone}" if u.phone else ""
            departments = ",".join(dept_org_map.get(dept_id, "") for dept_id in user_departments_map.get(u.id, []))
            leaders = ",".join(user_username_map.get(leader_id, "") for leader_id in user_leaders_map.get(u.id, []))

            base_info = [u.username, u.full_name, u.email, phone, departments, leaders]
            self.writer.add_row(base_info, u.extras or {})

        return self.writer.get_workbook()

    def _build_dept_org_map(self) -> Dict[int, str]:
        """
        获取部门与组织关系的映射表

        :returns: {dept_id: organization} 例如：{1: "总公司", 2: "总公司/深圳总部"}
        """
        dept_name_map = dict(
            DataSourceDepartment.objects.filter(data_source=self.data_source).values_list("id", "name")
        )
        relations = DataSourceDepartmentRelation.objects.filter(data_source=self.data_source)

        dept_org_map = {}

        def _build_by_recursive(rel: DataSourceDepartmentRelation, parent_org: str):
            dept_id = int(rel.department_id)
            dept_name = dept_name_map[dept_id]

            current_org = "/".join([parent_org, dept_name]) if parent_org else dept_name
            dept_org_map[dept_id] = current_org

            for child in rel.get_children():
                _build_by_recursive(child, current_org)

        # 使用 cached_tree 避免在后续使用 get_children 时候触发 DB 查询
        # 注：get_ascendants 无法使用 mptt 自带的缓存，暂不考虑在查询部门组织信息时使用
        for rel in relations.get_cached_trees():
            _build_by_recursive(rel, "")

        return dept_org_map

    def _build_user_departments_map(self) -> Dict[int, List[int]]:
        """
        获取用户与部门关系的映射表

        :returns: {user_id: [dept_id1, dept_id2, ...]}
        """
        relations = (
            DataSourceDepartmentUserRelation.objects.filter(user__in=self.users)
            .order_by("user_id")
            .values("user_id", "department_id")
        )
        return {
            user_id: sorted([r["department_id"] for r in group])
            for user_id, group in groupby(relations, key=lambda r: r["user_id"])
        }

    def _build_user_leaders_map(self) -> Dict[int, List[int]]:
        """
        获取用户与 leader 关系的映射表

        :returns: {user_id: [leader_id1, leader_id2, ...]}
        """
        relations = (
            DataSourceUserLeaderRelation.objects.filter(user__in=self.users)
            .order_by("user_id")
            .values("user_id", "leader_id")
        )
        return {
            user_id: sorted([r["leader_id"] for r in group])
            for user_id, group in groupby(relations, key=lambda r: r["user_id"])
        }

    def _build_user_username_map(self) -> Dict[int, str]:
        """获取用户与用户名的映射表"""
        return dict(self.users.values_list("id", "username"))
